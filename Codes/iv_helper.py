import os
import sys
import time
import numpy
import xlsxwriter as xw
from tqdm import tqdm
from openpyxl import load_workbook


time_start = time.process_time()

# 1.规划汇总数据文件名
path = input('请输入路径(将文件管理器中的路径复制过来即可):')
os.chdir(path)
first_folder = str(os.getcwd())
index = first_folder.rfind('\\')
result_name = first_folder[index+1 :] + '.xlsx'


# 2.创建汇总数据工作簿
workbook = xw.Workbook(result_name)
jv_workbook = xw.Workbook('JV_curve_data.xlsx')
format = workbook.add_format({'align': 'center', 'font_color': 'red', 'bold': True})
format1 = workbook.add_format({'align': 'center'})
format2 = workbook.add_format({'align': 'center', 'font_color': 'blue'})
format5 = jv_workbook.add_format({'align': 'center', 'font_color': 'blue'})


# 3.读取并写入数据
points_data = []
all_points_data = []
find_P_max = []
single_data = []
data = []

folders_list = os.listdir() # 获取所有子文件夹
for i in folders_list:
    # 创建对应组名的工作表
    worksheet = workbook.add_worksheet(i)
    worksheet.set_column(0, 5, 15)
    worksheet.set_column(6, 7, 25)
    worksheet.write(0, 0, 'file name', format)
    worksheet.write(0, 1, 'Mode', format)
    worksheet.write(0, 2, 'Voc (V)', format)
    worksheet.write(0, 3, 'Jsc (mA/cm^2)', format)
    worksheet.write(0, 4, 'FF (%)', format)
    worksheet.write(0, 5, 'PCE (%)', format)
    worksheet.write(0, 6, 'Device area (cm^2)', format)
    worksheet.write(0, 7, 'Author: umaru', format)
    worksheet.write(1, 7, 'From', format)
    worksheet.write(2, 7, '1 kWh Group', format)
    worksheet.write(3, 7, 'Funsom - Soochow.U', format)
    worksheet1 = jv_workbook.add_worksheet(i)
    os.chdir(i)
    files_list = os.listdir()
    for j in files_list:
        lines = 1
        file_to_read = open(j, 'r')
        single_data.append(j)
        points_data.append(j)
        while True:
            if len(single_data) == 5:
                try:
                    P_max = max(find_P_max)
                    FF = P_max / (abs(single_data[3]) * abs(single_data[4]))
                    PCE = abs(single_data[3]) * abs(single_data[4]) * FF
                    single_data.append(FF)
                    single_data.append(PCE)
                    data.append(single_data)
                    all_points_data.append(points_data)
                    single_data = []
                    points_data = []
                    find_P_max = []
                    break
                except:
                    single_data = []
                    points_data = []
                    break
            if lines == 3: # 通过检测填充数据判断正反扫
                line = file_to_read.readline()
                line = line.strip('\n')
                line = line.strip('Fill Factor = ')
                if eval(line) >= 1:
                    single_data.append('Forward')
                    points_data.append('Forward')
                else:
                    single_data.append('Reverse')
                    points_data.append('Reverse')
                lines += 1
            if lines == 7:
                line = file_to_read.readline()
                line = line.strip('\n')
                line = line.strip('Device area (cm^2) = ')
                single_data.append(eval(line))
                lines += 1
            if lines == 8: # 识别空扫数据
                line = file_to_read.readline()
                line = line.strip('\n')
                line = line.strip('Isd (A) = ')
                if eval(line) == 0:
                    single_data = []
                    break
                else:
                    lines += 1
                    continue
            if lines >= 11:
                try:
                    line = file_to_read.readline()
                    if line != '':
                        line = line.strip('\n').split('\t')
                        points_data.append(line)
                        lines += 1
                    else:
                        # 计算正确 Voc
                        for m in points_data[2:]:
                            index1 = points_data.index(m)
                            next_point = points_data[index1 + 1]
                            if numpy.sign(eval(m[1])) == numpy.sign(eval(next_point[1])):
                                continue
                            else:
                                correct_Voc = eval(m[0]) - eval(m[1]) * ((eval(next_point[0]) - eval(m[0])) / (eval(next_point[1]) - eval(m[1])))
                                single_data.append(correct_Voc)
                                break
                        # 计算正确 Jsc
                        for m in points_data[2:]:
                            index2 = points_data.index(m)
                            next_point = points_data[index2 + 1]
                            if numpy.sign(eval(m[0])) == numpy.sign(eval(next_point[0])):
                                continue
                            else:
                                correct_Jsc = eval(m[1]) - eval(m[0]) * ((eval(next_point[1]) - eval(m[1])) / (eval(next_point[0]) - eval(m[0])))
                                single_data.append(correct_Jsc)
                                break
                        # 寻找 P_max
                        for m in points_data[2:]:
                            if numpy.sign(eval(m[0])) == numpy.sign(single_data[3]) and numpy.sign(eval(m[1])) == numpy.sign(single_data[4]):
                                is_P_max = abs(eval(m[0])) * abs(eval(m[1]))
                                find_P_max.append(is_P_max)
                except:
                    single_data = []
                    points_data = []
                    break
            else:
                line = file_to_read.readline()
                lines += 1
                continue

    data.sort(key=lambda x:x[6], reverse=True) # 根据 PCE 对数据进行降序排列

    # 将整理好的数据写入对应工作表
    row_num = 1
    for m in data:
        if row_num <= len(data):
            # 写入文件名
            worksheet.write(row_num, 0, m[0], format1)
            # 写入扫描模式
            if m[1] == 'Forward':
                worksheet.write(row_num, 1, m[1], format2)
            else:
                worksheet.write(row_num, 1, m[1], format1)
            # 写入 Voc
            worksheet.write(row_num, 2, abs(round(m[3], 3)), format1)
            # 写入 Jsc
            worksheet.write(row_num, 3, abs(round(m[4], 3)), format1)
            # 写入 FF
            worksheet.write(row_num, 4, round(m[5], 3), format1)
            # 写入 PCE
            worksheet.write(row_num, 5, round(m[6], 3), format1)
            # 写入有效面积
            worksheet.write(row_num, 6, m[2], format1)
            row_num += 1

    # 将 jv 的数据写入对应工作表
    column_num1 = 0
    for n in all_points_data:
        if n[1] == 'Forward':
            worksheet1.write(0, column_num1, n[0], format5)
        else:
            worksheet1.write(0, column_num1, n[0])
        row_num1 = 1
        for n1 in n[2:]:
            if row_num1 <= len(n[2:]):
                worksheet1.write(row_num1, column_num1, eval(n1[0]))
                worksheet1.write(row_num1, column_num1+1, eval(n1[1]))
                row_num1 += 1
        column_num1 += 3
    all_points_data = []
    data = []
    os.chdir(first_folder)

workbook.close()
jv_workbook.close()


# 5.生成统计文件 -- 方便绘制统计图
target_workbook = load_workbook(result_name)

Statistical_Voc = []
Statistical_Jsc = []
Statistical_FF = []
Statistical_PCE = []
for sheet in target_workbook.worksheets:
    sheet_name = str(sheet).split('"')[1]
    row_num = sheet.max_row
    Temple_Voc, Temple_Jsc, Temple_FF, Temple_PCE = [sheet_name], [sheet_name], [sheet_name], [sheet_name]
    for row in range(2, row_num+2):
        if sheet.cell(row, 3).value != None:
            Temple_Voc.append([sheet.cell(row, 3).value, sheet.cell(row, 2).value])
            Temple_Jsc.append([sheet.cell(row, 4).value, sheet.cell(row, 2).value])
            Temple_FF.append([sheet.cell(row, 5).value, sheet.cell(row, 2).value])
            Temple_PCE.append([sheet.cell(row, 6).value, sheet.cell(row, 2).value])
        else:
            Statistical_Voc.append(Temple_Voc)
            Statistical_Jsc.append(Temple_Jsc)
            Statistical_FF.append(Temple_FF)
            Statistical_PCE.append(Temple_PCE)
            break

Statistical_workbook = xw.Workbook('Statistical data.xlsx')
format3 = Statistical_workbook.add_format({'align': 'center'})
format4 = Statistical_workbook.add_format({'align': 'center', 'font_color': 'blue'})
Voc_sheet = Statistical_workbook.add_worksheet('Voc')
Jsc_sheet = Statistical_workbook.add_worksheet('Jsc')
FF_sheet = Statistical_workbook.add_worksheet('FF')
PCE_sheet = Statistical_workbook.add_worksheet('PCE')

def write_data(data_list, target_sheet):
    c_num = 0
    columns_num = len(data_list)
    target_sheet.set_column(0, columns_num-1, 15)
    for data in data_list:
        if c_num <= len(data_list):
            r_num = 0
            for i in data:
                if r_num <= len(data)-1:
                    if type(i) == str:
                        target_sheet.write(r_num, c_num, i, format3)
                    else:
                        if i[1] == 'Forward':
                            target_sheet.write(r_num, c_num, i[0], format4)
                        else:
                            target_sheet.write(r_num, c_num, i[0], format3)
                    r_num += 1
            c_num += 1

write_data(Statistical_Voc, Voc_sheet)
write_data(Statistical_Jsc, Jsc_sheet)
write_data(Statistical_FF, FF_sheet)
write_data(Statistical_PCE, PCE_sheet)

Statistical_workbook.close()

time_end = time.process_time()
time_sum = round((time_end - time_start), 2)
counter = int(time_sum / 0.01)

for k in tqdm(range(counter)):
    time.sleep(0.01)

print('数据处理完毕...')

imported_modules_list = []
for m in sys.modules:
    imported_modules_list.append(m)

if 'menu' in imported_modules_list:
    del sys.modules['menu']
    import menu
else:
    import menu